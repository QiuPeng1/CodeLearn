
import openpyxl

#file = '中奖码2.xlsx'
uinput = input("输入文件名：")
workbook = openpyxl.load_workbook(uinput)
sheet_names = workbook.sheetnames
sheet1 = workbook[sheet_names[0]]
nrows = sheet1.rows

print(nrows)

#读取Excel sheet1中的所有数据
allDatas = []
level1 = 0
level2 = 0
level3 = 0
level4 = 0
level5 = 0
level6 = 0
level7 = 0
level8 = 0
level9 = 0

for row in nrows:
    lines = [cell.value for cell in row]
    allDatas.append(lines)
    if lines[1] == 1:
        level1 = level1 + 1
    elif lines[1] == 2:
        level2 = level2 + 1
    elif lines[1] == 3:
        level3 = level3 + 1
    elif lines[1] == 4:
        level4 = level4 + 1
    elif lines[1] == 5:
        level5 = level5 + 1
    elif lines[1] == 6:
        level6 = level6 + 1
    elif lines[1] ==7:
        level7 = level7 + 1
    elif lines[1] == 8:
        level8 = level8 + 1
    elif lines[1] == 9:
        level9 = level9 + 1

level1_cnt = level1/50
level2_cnt = level2/50
level3_cnt = level3/50
level4_cnt = level4/50
level5_cnt = level5/50
level6_cnt = level6/50
level7_cnt = level7/50
level8_cnt = level8/50
level9_cnt = level9/50
L1 = [level1, level2, level3, level4, level5, level6, level7, level8, level9]
L2 = [level1_cnt,level1_cnt, level2_cnt, level3_cnt, level4_cnt, level5_cnt, level6_cnt, level7_cnt, level8_cnt, level9_cnt]
print(L1)
print(L2)

#划分班级
title = []
class1 = []
class2 = []
class3 = []
class4 = []
class5 = []
class6 = []
class7 = []
class8 = []
class9 = []
class10 = []
class11 = []
class12 = []
class13 = []
class14 = []
class15 = []
class16 = []
class17 = []
class18 = []
class19 = []
class20 = []
class21 = []
class22 = []
class23 = []
class24 = []
class25 = []
class26 = []
class27 = []
class28 = []
class29 = []
class30 = []
class31 = []
class32 = []
class33 = []
class34 = []
class35 = []
class36 = []
class37 = []
class38 = []
class39 = []
class40 = []
class41 = []
class42 = []
class43 = []
class44 = []
class45 = []
class46 = []
class47 = []
class48 = []
class49 = []
class50 = []

userrow = 0
for i in allDatas:
     if i[1] == 1 or i[1] == 2 or i[1] == 3 or i[1] == 4 or i[1] == 5 or i[1] == 6 or i[1] == 7 or i[1] == 8 or i[1] == 9:
         userrow = userrow + 1
         if userrow < L2[i[1]]+1:
             class1.append(i)
         elif userrow < L2[i[1]]*2+1:
             class2.append(i)
         elif userrow < L2[i[1]]*3+1:
             class3.append(i)
         elif userrow < L2[i[1]]*4+1:
             class4.append(i)
         elif userrow < L2[i[1]]*5+1:
             class5.append(i)
         elif userrow < L2[i[1]]*6+1:
             class6.append(i)
         elif userrow < L2[i[1]]*7+1:
             class7.append(i)
         elif userrow < L2[i[1]]*8+1:
             class8.append(i)
         elif userrow < L2[i[1]]*9+1:
             class9.append(i)
         elif userrow < L2[i[1]]*10+1:
             class10.append(i)
         elif userrow < L2[i[1]]*11+1:
             class11.append(i)
         elif userrow < L2[i[1]]*12+1:
             class12.append(i)
         elif userrow < L2[i[1]]*13+1:
             class13.append(i)
         elif userrow < L2[i[1]]*14+1:
             class14.append(i)
         elif userrow < L2[i[1]]*15+1:
             class15.append(i)
         elif userrow < L2[i[1]]*16+1:
             class16.append(i)
         elif userrow < L2[i[1]]*17+1:
             class17.append(i)
         elif userrow < L2[i[1]]*18+1:
             class18.append(i)
         elif userrow < L2[i[1]]*19+1:
             class19.append(i)
         elif userrow < L2[i[1]]*20+1:
             class20.append(i)
         elif userrow < L2[i[1]]*21+1:
             class21.append(i)
         elif userrow < L2[i[1]]*22+1:
             class22.append(i)
         elif userrow < L2[i[1]]*23+1:
             class23.append(i)
         elif userrow < L2[i[1]]*24+1:
             class24.append(i)
         elif userrow < L2[i[1]]*25+1:
             class25.append(i)
         elif userrow < L2[i[1]]*26+1:
             class26.append(i)
         elif userrow < L2[i[1]]*27+1:
             class27.append(i)
         elif userrow < L2[i[1]]*28+1:
             class28.append(i)
         elif userrow < L2[i[1]]*29+1:
             class29.append(i)
         elif userrow < L2[i[1]]*30+1:
             class30.append(i)
         elif userrow < L2[i[1]]*31+1:
             class31.append(i)
         elif userrow < L2[i[1]]*32+1:
             class32.append(i)
         elif userrow < L2[i[1]]*33+1:
             class33.append(i)
         elif userrow < L2[i[1]]*34+1:
             class34.append(i)
         elif userrow < L2[i[1]]*35+1:
             class35.append(i)
         elif userrow < L2[i[1]]*36+1:
             class36.append(i)
         elif userrow < L2[i[1]]*37+1:
             class37.append(i)
         elif userrow < L2[i[1]]*38+1:
             class38.append(i)
         elif userrow < L2[i[1]]*39+1:
             class39.append(i)
         elif userrow < L2[i[1]]*40+1:
             class40.append(i)
         elif userrow < L2[i[1]]*41+1:
             class41.append(i)
         elif userrow < L2[i[1]]*42+1:
             class42.append(i)
         elif userrow < L2[i[1]]*43+1:
             class43.append(i)
         elif userrow < L2[i[1]]*44+1:
             class44.append(i)
         elif userrow < L2[i[1]]*45+1:
             class45.append(i)
         elif userrow < L2[i[1]]*46+1:
             class46.append(i)
         elif userrow < L2[i[1]]*47+1:
             class47.append(i)
         elif userrow < L2[i[1]]*48+1:
             class48.append(i)
         elif userrow < L2[i[1]]*49+1:
             class49.append(i)
         elif userrow < L2[i[1]]*50:
             class50.append(i)
         elif userrow == L2[i[1]]*50:
             class50.append(i)
             userrow = 0
         else:
             userrow = 0
     else:
         title.append(i)
class1.insert(0,title[0]) # 把每个生成的Excel中插入title
class2.insert(0,title[0])
class3.insert(0,title[0])
class4.insert(0,title[0])
class5.insert(0,title[0])
class6.insert(0,title[0])
class7.insert(0,title[0])
class8.insert(0,title[0])
class9.insert(0,title[0])
class10.insert(0,title[0])
class11.insert(0,title[0]) # 把每个生成的Excel中插入title
class12.insert(0,title[0])
class13.insert(0,title[0])
class14.insert(0,title[0])
class15.insert(0,title[0])
class16.insert(0,title[0])
class17.insert(0,title[0])
class18.insert(0,title[0])
class19.insert(0,title[0])
class20.insert(0,title[0])
class21.insert(0,title[0]) # 把每个生成的Excel中插入title
class22.insert(0,title[0])
class23.insert(0,title[0])
class24.insert(0,title[0])
class25.insert(0,title[0])
class26.insert(0,title[0])
class27.insert(0,title[0])
class28.insert(0,title[0])
class29.insert(0,title[0])
class30.insert(0,title[0])
class31.insert(0,title[0]) # 把每个生成的Excel中插入title
class32.insert(0,title[0])
class33.insert(0,title[0])
class34.insert(0,title[0])
class35.insert(0,title[0])
class36.insert(0,title[0])
class37.insert(0,title[0])
class38.insert(0,title[0])
class39.insert(0,title[0])
class40.insert(0,title[0])
class41.insert(0,title[0]) # 把每个生成的Excel中插入title
class42.insert(0,title[0])
class43.insert(0,title[0])
class44.insert(0,title[0])
class45.insert(0,title[0])
class46.insert(0,title[0])
class47.insert(0,title[0])
class48.insert(0,title[0])
class49.insert(0,title[0])
class50.insert(0,title[0])

#print(class1)
# 把每个年级的数据分别保存为一个文件
def get_datas(datas,path):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
# worksheet.title = "Class1_Datas"
    counter = 0
    for lines in datas:
        #print(lines)
        counter = counter + 1
        for i in range(len(lines)):
            worksheet.cell(counter, i + 1, lines[i])
        workbook.save(path)
    return workbook
 
if __name__ == "__main__":
    #设定生成文件的路径并指明文件名
    class1_datas = R"class1_datas.xlsx"
    class2_datas = R"class2_datas.xlsx"
    class3_datas = R"class3_datas.xlsx"
    class4_datas = R"class4_datas.xlsx"
    class5_datas = R"class5_datas.xlsx"
    class6_datas = R"class6_datas.xlsx"
    class7_datas = R"class7_datas.xlsx"
    class8_datas = R"class8_datas.xlsx"
    class9_datas = R"class9_datas.xlsx"
    class10_datas = R"class10_datas.xlsx"
    class11_datas = R"class11_datas.xlsx"
    class12_datas = R"class12_datas.xlsx"
    class13_datas = R"class13_datas.xlsx"
    class14_datas = R"class14_datas.xlsx"
    class15_datas = R"class15_datas.xlsx"
    class16_datas = R"class16_datas.xlsx"
    class17_datas = R"class17_datas.xlsx"
    class18_datas = R"class18_datas.xlsx"
    class19_datas = R"class19_datas.xlsx"
    class20_datas = R"class20_datas.xlsx"  
    class21_datas = R"class21_datas.xlsx"
    class22_datas = R"class22_datas.xlsx"
    class23_datas = R"class23_datas.xlsx"
    class24_datas = R"class24_datas.xlsx"
    class25_datas = R"class25_datas.xlsx"
    class26_datas = R"class26_datas.xlsx"
    class27_datas = R"class27_datas.xlsx"
    class28_datas = R"class28_datas.xlsx"
    class29_datas = R"class29_datas.xlsx"
    class30_datas = R"class30_datas.xlsx"
    class31_datas = R"class31_datas.xlsx"
    class32_datas = R"class32_datas.xlsx"
    class33_datas = R"class33_datas.xlsx"
    class34_datas = R"class34_datas.xlsx"
    class35_datas = R"class35_datas.xlsx"
    class36_datas = R"class36_datas.xlsx"
    class37_datas = R"class37_datas.xlsx"
    class38_datas = R"class38_datas.xlsx"
    class39_datas = R"class39_datas.xlsx"
    class40_datas = R"class40_datas.xlsx"
    class41_datas = R"class41_datas.xlsx"
    class42_datas = R"class42_datas.xlsx"
    class43_datas = R"class43_datas.xlsx"
    class44_datas = R"class44_datas.xlsx"
    class45_datas = R"class45_datas.xlsx"
    class46_datas = R"class46_datas.xlsx"
    class47_datas = R"class47_datas.xlsx"
    class48_datas = R"class48_datas.xlsx"
    class49_datas = R"class49_datas.xlsx"
    class50_datas = R"class50_datas.xlsx"
 
    #调用方法写入文件
    print("中奖单1生成成功！",get_datas(class1,class1_datas))
    print("中奖单2生成成功！",get_datas(class2, class2_datas))
    print("中奖单3生成成功！", get_datas(class3, class3_datas))
    print("中奖单4生成成功！",get_datas(class4, class4_datas))
    print("中奖单5生成成功！",get_datas(class5, class5_datas))
    print("中奖单6生成成功！",get_datas(class6,class6_datas))
    print("中奖单7生成成功！",get_datas(class7, class7_datas))
    print("中奖单8生成成功！", get_datas(class8, class8_datas))
    print("中奖单9生成成功！",get_datas(class9, class9_datas))
    print("中奖单10生成成功！",get_datas(class10, class10_datas))
    print("中奖单11生成成功！",get_datas(class11,class11_datas))
    print("中奖单12生成成功！",get_datas(class12, class12_datas))
    print("中奖单13生成成功！", get_datas(class13, class13_datas))
    print("中奖单14生成成功！",get_datas(class14, class14_datas))
    print("中奖单15生成成功！",get_datas(class15, class15_datas))
    print("中奖单16生成成功！",get_datas(class16,class16_datas))
    print("中奖单17生成成功！",get_datas(class17, class17_datas))
    print("中奖单18生成成功！", get_datas(class18, class18_datas))
    print("中奖单19生成成功！",get_datas(class19, class19_datas))
    print("中奖单20生成成功！",get_datas(class20, class20_datas))
    print("中奖单21生成成功！",get_datas(class21,class21_datas))
    print("中奖单22生成成功！",get_datas(class22, class22_datas))
    print("中奖单23生成成功！", get_datas(class23, class23_datas))
    print("中奖单24生成成功！",get_datas(class24, class24_datas))
    print("中奖单25生成成功！",get_datas(class25, class25_datas))
    print("中奖单26生成成功！",get_datas(class26,class26_datas))
    print("中奖单27生成成功！",get_datas(class27, class27_datas))
    print("中奖单28生成成功！", get_datas(class28, class28_datas))
    print("中奖单29生成成功！",get_datas(class29, class29_datas))
    print("中奖单30生成成功！",get_datas(class30, class30_datas))
    print("中奖单31生成成功！",get_datas(class31,class31_datas))
    print("中奖单32生成成功！",get_datas(class32, class32_datas))
    print("中奖单33生成成功！", get_datas(class33, class33_datas))
    print("中奖单34生成成功！",get_datas(class34, class34_datas))
    print("中奖单35生成成功！",get_datas(class35, class35_datas))
    print("中奖单36生成成功！",get_datas(class36,class36_datas))
    print("中奖单37生成成功！",get_datas(class37, class37_datas))
    print("中奖单38生成成功！", get_datas(class38, class38_datas))
    print("中奖单39生成成功！",get_datas(class39, class39_datas))
    print("中奖单40生成成功！",get_datas(class40, class40_datas))
    print("中奖单41生成成功！",get_datas(class41,class41_datas))
    print("中奖单42生成成功！",get_datas(class42, class42_datas))
    print("中奖单43生成成功！", get_datas(class43, class43_datas))
    print("中奖单44生成成功！",get_datas(class44, class44_datas))
    print("中奖单45生成成功！",get_datas(class45, class45_datas))
    print("中奖单46生成成功！",get_datas(class46,class46_datas))
    print("中奖单47生成成功！",get_datas(class47, class47_datas))
    print("中奖单48生成成功！", get_datas(class48, class48_datas))
    print("中奖单49生成成功！",get_datas(class49, class49_datas))
    print("中奖单50生成成功！",get_datas(class50, class50_datas))